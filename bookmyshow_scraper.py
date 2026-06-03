"""
BookMyShow Movie Show Listings Scraper
A Playwright-based web scraper that collects movie show listings from BookMyShow
with manual login support and ethical scraping practices.
"""

import time
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path
from playwright.sync_api import sync_playwright, Page, Browser, BrowserContext
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class BookMyShowScraper:
    """Scraper for BookMyShow movie show listings."""
    
    # Wait times (in seconds) for ethical scraping
    SHORT_WAIT = 2
    MEDIUM_WAIT = 5
    LONG_WAIT = 10
    
    # Selectors for key elements
    SELECTORS = {
        'movie_cards': '[data-testid="movieCard"]',
        'theatre_name': '[data-testid="theatreName"]',
        'show_time': '[data-testid="showTime"]',
        'availability': '[data-testid="availability"]',
        'price': '[data-testid="price"]',
        'book_button': 'button:has-text("Book")',
    }
    
    def __init__(self, headless: bool = False):
        """
        Initialize the scraper.
        
        Args:
            headless: Run browser in headless mode (default: False for visible browser)
        """
        self.headless = headless
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.collected_shows = []
        self.current_city = None
        self.current_movie = None
        self.current_date = None
        
    def start(self):
        """Start the Playwright browser."""
        try:
            self.playwright = sync_playwright().start()
            self.browser = self.playwright.chromium.launch(
                headless=self.headless,
                args=['--disable-blink-features=AutomationControlled']
            )
            self.context = self.browser.new_context(
                viewport={'width': 1920, 'height': 1080}
            )
            self.page = self.context.new_page()
            logger.info("Browser started successfully")
        except Exception as e:
            logger.error(f"Failed to start browser: {e}")
            raise
    
    def stop(self):
        """Stop the Playwright browser."""
        try:
            if self.page:
                self.page.close()
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            if self.playwright:
                self.playwright.stop()
            logger.info("Browser stopped successfully")
        except Exception as e:
            logger.error(f"Error stopping browser: {e}")
    
    def open_bookmyshow(self):
        """Open BookMyShow website."""
        try:
            logger.info("Opening BookMyShow website...")
            self.page.goto('https://www.bookmyshow.com/', wait_until='networkidle', timeout=30000)
            self.page.wait_for_load_state('domcontentloaded')
            time.sleep(self.MEDIUM_WAIT)
            logger.info("BookMyShow loaded successfully")
        except Exception as e:
            logger.error(f"Failed to open BookMyShow: {e}")
            raise
    
    def wait_for_manual_login(self, timeout_seconds: int = 300):
        """
        Wait for user to manually login.
        
        Args:
            timeout_seconds: Maximum time to wait for login (default: 5 minutes)
        """
        logger.info(f"Waiting for manual login (up to {timeout_seconds} seconds)...")
        logger.info("Please login and select your city in the browser window")
        
        try:
            # Wait for a navigation or significant page change after login
            # Look for elements that typically appear after login
            self.page.wait_for_load_state('networkidle', timeout=timeout_seconds * 1000)
            time.sleep(self.SHORT_WAIT)
            logger.info("Login detected, proceeding...")
        except Exception as e:
            logger.warning(f"Timeout waiting for login: {e}")
    
    def detect_city(self) -> Optional[str]:
        """
        Detect the currently selected city.
        
        Returns:
            City name if detected, None otherwise
        """
        try:
            # Try to find city selector/display
            city_element = self.page.locator(
                'button:has-text("Change"), [data-testid="citySelector"], .citySelector'
            ).first
            
            if city_element.is_visible():
                city_text = city_element.text_content()
                if city_text:
                    self.current_city = city_text.strip()
                    logger.info(f"Detected city: {self.current_city}")
                    return self.current_city
        except Exception as e:
            logger.warning(f"Could not detect city: {e}")
        
        return None
    
    def scroll_to_load_more(self, scroll_times: int = 3):
        """
        Scroll down to load more content.
        
        Args:
            scroll_times: Number of times to scroll
        """
        try:
            for i in range(scroll_times):
                logger.info(f"Scrolling... ({i+1}/{scroll_times})")
                self.page.evaluate('window.scrollBy(0, window.innerHeight)')
                time.sleep(self.SHORT_WAIT)
        except Exception as e:
            logger.warning(f"Error during scrolling: {e}")
    
    def extract_show_listings(self) -> List[Dict]:
        """
        Extract visible movie show listings from the current page.
        
        Returns:
            List of dictionaries containing show information
        """
        shows = []
        
        try:
            logger.info("Extracting visible show listings...")
            
            # Get all movie cards visible on page
            movie_cards = self.page.locator('[data-testid="movieCard"]')
            card_count = movie_cards.count()
            
            logger.info(f"Found {card_count} movie cards")
            
            for idx in range(min(card_count, 50)):  # Limit to prevent excessive processing
                try:
                    card = movie_cards.nth(idx)
                    
                    # Check if card is in viewport
                    if not card.is_visible():
                        continue
                    
                    # Extract movie name
                    movie_name = self._safe_extract_text(
                        card.locator('[data-testid="movieTitle"]')
                    )
                    
                    if not movie_name:
                        continue
                    
                    self.current_movie = movie_name
                    
                    # Extract shows within this movie card
                    show_elements = card.locator('[data-testid="showDetails"]')
                    show_count = show_elements.count()
                    
                    for show_idx in range(show_count):
                        try:
                            show = show_elements.nth(show_idx)
                            
                            show_data = {
                                'movie': movie_name,
                                'city': self.current_city or 'Not Detected',
                                'date': self._extract_date(),
                                'theatre_name': self._safe_extract_text(
                                    show.locator('[data-testid="theatreName"]')
                                ),
                                'show_time': self._safe_extract_text(
                                    show.locator('[data-testid="showTime"]')
                                ),
                                'availability': self._safe_extract_text(
                                    show.locator('[data-testid="availability"]')
                                ),
                                'price': self._safe_extract_text(
                                    show.locator('[data-testid="price"]')
                                ),
                                'scraped_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            
                            # Only add if we have essential data
                            if show_data['theatre_name'] and show_data['show_time']:
                                shows.append(show_data)
                                logger.debug(f"Extracted show: {show_data['theatre_name']} - {show_data['show_time']}")
                        
                        except Exception as e:
                            logger.debug(f"Error extracting individual show: {e}")
                            continue
                
                except Exception as e:
                    logger.debug(f"Error processing card {idx}: {e}")
                    continue
            
            logger.info(f"Successfully extracted {len(shows)} show listings")
            self.collected_shows.extend(shows)
            return shows
        
        except Exception as e:
            logger.error(f"Error extracting show listings: {e}")
            return []
    
    def _safe_extract_text(self, locator) -> str:
        """
        Safely extract text from a locator.
        
        Args:
            locator: Playwright locator object
            
        Returns:
            Extracted text or empty string
        """
        try:
            if locator.count() > 0 and locator.first.is_visible():
                text = locator.first.text_content()
                return text.strip() if text else ''
        except Exception:
            pass
        return ''
    
    def _extract_date(self) -> str:
        """
        Extract the current date being viewed.
        
        Returns:
            Date string or empty string
        """
        try:
            # Try to find active date element
            date_locator = self.page.locator('[data-testid="dateSelector"] .active, [data-testid="selectedDate"]')
            if date_locator.count() > 0:
                return date_locator.first.text_content().strip()
        except Exception:
            pass
        
        # Return today's date as fallback
        return datetime.now().strftime('%Y-%m-%d')
    
    def save_to_excel(self, filename: str = 'bookmyshow_listings.xlsx'):
        """
        Save collected show listings to an Excel file.
        
        Args:
            filename: Name of the output Excel file
        """
        try:
            if not self.collected_shows:
                logger.warning("No shows to save")
                return
            
            logger.info(f"Saving {len(self.collected_shows)} shows to {filename}...")
            
            # Create DataFrame
            df = pd.DataFrame(self.collected_shows)
            
            # Reorder columns for better readability
            columns_order = [
                'movie', 'city', 'date', 'theatre_name', 'show_time',
                'availability', 'price', 'scraped_at'
            ]
            df = df[[col for col in columns_order if col in df.columns]]
            
            # Save to Excel with formatting
            filepath = Path(filename)
            df.to_excel(filepath, index=False, sheet_name='Shows')
            
            # Format the Excel file
            self._format_excel(filepath)
            
            logger.info(f"Successfully saved to {filepath.absolute()}")
            print(f"\n✓ Data saved to: {filepath.absolute()}")
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")
            raise
    
    def _format_excel(self, filepath: Path):
        """
        Apply formatting to the Excel file.
        
        Args:
            filepath: Path to the Excel file
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Header formatting
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Column width adjustment
            column_widths = {
                'A': 25,  # movie
                'B': 15,  # city
                'C': 12,  # date
                'D': 25,  # theatre_name
                'E': 12,  # show_time
                'F': 20,  # availability
                'G': 12,  # price
                'H': 20,  # scraped_at
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Data cell formatting
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            wb.save(filepath)
            logger.info("Excel file formatted successfully")
        
        except Exception as e:
            logger.warning(f"Could not format Excel file: {e}")
    
    def run(self, output_filename: str = 'bookmyshow_listings.xlsx'):
        """
        Run the complete scraping workflow.
        
        Args:
            output_filename: Name of the output Excel file
        """
        try:
            # Start browser
            self.start()
            
            # Open BookMyShow
            self.open_bookmyshow()
            
            # Wait for manual login
            self.wait_for_manual_login()
            
            # Detect city
            self.detect_city()
            
            # Scroll to load more content
            logger.info("Loading more shows by scrolling...")
            self.scroll_to_load_more(scroll_times=5)
            
            # Extract show listings
            self.extract_show_listings()
            
            # Save to Excel
            self.save_to_excel(output_filename)
            
            # Keep browser open for review
            logger.info("Scraping complete. Browser will remain open for 30 seconds.")
            time.sleep(30)
            
        except Exception as e:
            logger.error(f"Error during scraping: {e}")
            raise
        
        finally:
            self.stop()


def main():
    """Main entry point."""
    
    print("=" * 60)
    print("BookMyShow Movie Show Listings Scraper")
    print("=" * 60)
    print("\nThis script will:")
    print("1. Open BookMyShow in a visible browser")
    print("2. Wait for you to manually login and select a city")
    print("3. Collect visible show listings")
    print("4. Save the data to an Excel file")
    print("\nNote: Please keep the browser window active and do not close it")
    print("      until the script completes.")
    print("\n" + "=" * 60 + "\n")
    
    try:
        # Initialize scraper
        scraper = BookMyShowScraper(headless=False)
        
        # Run scraping
        scraper.run()
        
        print("\n✓ Scraping completed successfully!")
        print("  Check 'bookmyshow_listings.xlsx' for the results.")
        
    except KeyboardInterrupt:
        print("\n\n✗ Scraping interrupted by user")
    except Exception as e:
        print(f"\n✗ Scraping failed: {e}")
        logger.exception("Full error trace:")


if __name__ == '__main__':
    main()
